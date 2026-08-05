import unittest
from io import BytesIO

from openpyxl import Workbook

from app.models.schemas import ProjectRecord
from app.services.spreadsheet_import_service import import_calibration_workbook


INFLUENT_HEADERS = [
    "工厂编号",
    "采样时间",
    "流量（立方米/日）",
    "化学需氧量（毫克/升）",
    "五日生化需氧量（毫克/升）",
    "氨氮（毫克/升）",
    "总氮（毫克/升）",
    "总磷（毫克/升）",
    "悬浮物（毫克/升）",
    "酸碱度",
    "水温（摄氏度）",
    "碱度（毫克/升，以碳酸钙计）",
]
EFFLUENT_HEADERS = [
    "工厂编号",
    "采样时间",
    "化学需氧量（毫克/升）",
    "氨氮（毫克/升）",
    "总氮（毫克/升）",
    "总磷（毫克/升）",
    "悬浮物（毫克/升）",
]
OPERATION_HEADERS = [
    "工厂编号",
    "记录时间",
    "水力停留时间（小时）",
    "污泥龄（日）",
    "内回流比",
    "污泥回流比",
    "好氧池溶解氧（毫克/升）",
    "曝气功率（千瓦）",
    "碱度（毫克/升，以碳酸钙计）",
    "二沉池固体捕集率",
    "生化池总有效容积（立方米）",
    "实际排泥流量（立方米/日）",
    "池内污泥浓度（毫克/升）",
    "回流污泥浓度（毫克/升）",
    "排泥污泥浓度（毫克/升）",
    "二沉池总表面积（平方米）",
    "二沉池有效水深（米）",
]


def workbook_bytes(header_row: int = 1, duplicate_operation: bool = False) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    rows = {
        "进水数据": (
            INFLUENT_HEADERS,
            ["甲厂", "2026-07-01 08:00:00", 8000, 300, 150, 35, 48, 5, 200, 7.2, 20, 240],
        ),
        "出水数据": (
            EFFLUENT_HEADERS,
            ["甲厂", "2026-07-01 08:00:00", 35, 2, 12, 0.4, 8],
        ),
        "运行参数": (
            OPERATION_HEADERS,
            ["甲厂", "2026-07-01 08:00:00", 12, 15, 2, 0.8, 2.2, 20, 240, 0.98,
             4000, 26.67, 3000, 8000, 7500, 650, 4],
        ),
    }
    for name, (headers, values) in rows.items():
        sheet = workbook.create_sheet(name)
        for _ in range(header_row - 1):
            sheet.append([])
        sheet.append(headers)
        sheet.append(values)
        if name == "运行参数" and duplicate_operation:
            sheet.append([*values[:6], 3.0, *values[7:]])
    target = BytesIO()
    workbook.save(target)
    return target.getvalue()


class SpreadsheetImportTests(unittest.TestCase):
    def setUp(self) -> None:
        self.project = ProjectRecord(
            name="导入测试",
            plant_name="甲厂",
            process_type="AAO",
        )

    def test_accepts_first_row_headers_and_string_dates(self) -> None:
        result = import_calibration_workbook(self.project, workbook_bytes())
        self.assertEqual(result.imported_count, 1)
        self.assertEqual(result.skipped_count, 0)
        self.assertEqual(result.groups, ["甲厂"])
        self.assertEqual(result.samples[0].parameters.hrt_h, 12)
        self.assertEqual(result.samples[0].parameters.waste_sludge_flow_m3_d, 26.67)
        self.assertEqual(result.samples[0].parameters.clarifier_surface_area_m2, 650)
        self.assertEqual(
            result.samples[0].parameters.operating_data_source,
            "measured",
        )
        self.assertEqual(result.quality_score, 100)
        self.assertEqual(result.readiness, "补充后校准")
        self.assertEqual(result.duplicate_key_count, 0)

    def test_accepts_fourth_row_headers(self) -> None:
        result = import_calibration_workbook(
            self.project, workbook_bytes(header_row=4)
        )
        self.assertEqual(result.imported_count, 1)

    def test_rejects_ambiguous_operation_rows(self) -> None:
        result = import_calibration_workbook(
            self.project, workbook_bytes(duplicate_operation=True)
        )
        self.assertEqual(result.imported_count, 0)
        self.assertEqual(result.skipped_count, 1)
        self.assertEqual(result.readiness, "不可校准")
        self.assertEqual(result.duplicate_key_count, 1)
        self.assertTrue(any("多条运行记录" in item for item in result.warnings))
        self.assertTrue(any("重复键" in item for item in result.recommendations))

    def test_censored_effluent_is_not_used_as_exact_measurement(self) -> None:
        content = workbook_bytes()
        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(content))
        sheet = workbook["出水数据"]
        sheet.cell(row=1, column=3, value="化学需氧量检出限符号")
        sheet.cell(row=2, column=3, value="<")
        target = BytesIO()
        workbook.save(target)
        result = import_calibration_workbook(self.project, target.getvalue())
        self.assertIsNone(result.samples[0].measured.cod_mg_l)
        self.assertTrue(any("检出限符号" in item for item in result.warnings))


if __name__ == "__main__":
    unittest.main()
