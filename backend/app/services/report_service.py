from datetime import datetime
from pathlib import Path

from app.models.schemas import (
    ProjectRecord,
    ReportFormat,
    ReportRequest,
    ReportResult,
    SimulationResult,
)


REPORT_DIR = Path(__file__).resolve().parents[2] / "generated_reports"
REPORT_DIR.mkdir(parents=True, exist_ok=True)

INDICATOR_LABELS = {
    "cod_mg_l": "化学需氧量",
    "nh4_n_mg_l": "氨氮",
    "tn_mg_l": "总氮",
    "tp_mg_l": "总磷",
    "tss_mg_l": "悬浮物",
}


def _safe_filename(value: str) -> str:
    allowed = "".join(
        character
        for character in value.strip()
        if character.isalnum() or character in ("-", "_", " ")
    )
    return allowed[:60] or "污水工艺建模评估报告"


def _create_pdf(
    path: Path, title: str, project: ProjectRecord, simulation: SimulationResult
) -> None:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    font_path = Path(__file__).resolve().parents[1] / "assets" / "NotoSansCJKsc-VF.ttf"
    if not font_path.is_file():
        macos_fallback = Path("/System/Library/Fonts/Supplemental/Arial Unicode.ttf")
        if macos_fallback.is_file():
            font_path = macos_fallback
        else:
            raise RuntimeError("中文报告字体未安装，请重新构建后端镜像。")
    pdfmetrics.registerFont(TTFont("NotoSansSC", str(font_path)))
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ChineseTitle",
        parent=styles["Title"],
        fontName="NotoSansSC",
        fontSize=20,
        leading=28,
        textColor=colors.HexColor("#174f43"),
    )
    heading_style = ParagraphStyle(
        "ChineseHeading",
        parent=styles["Heading2"],
        fontName="NotoSansSC",
        fontSize=13,
        leading=20,
        textColor=colors.HexColor("#176b58"),
        spaceBefore=10,
    )
    body_style = ParagraphStyle(
        "ChineseBody",
        parent=styles["BodyText"],
        fontName="NotoSansSC",
        fontSize=9,
        leading=15,
    )
    doc = SimpleDocTemplate(
        str(path),
        pagesize=A4,
        rightMargin=18 * mm,
        leftMargin=18 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
        title=title,
        author=project.owner or "清澜智评",
    )
    story = [
        Paragraph(title, title_style),
        Spacer(1, 8 * mm),
        Paragraph("一、项目与模型概况", heading_style),
        Paragraph(
            f"项目：{project.name}<br/>污水厂：{project.plant_name}<br/>"
            f"主体工艺：{project.process_type.value}<br/>"
            f"计算引擎：{simulation.engine}<br/>"
            f"结果范围：{'生化段加强化处理最终出水' if simulation.advanced_treatment_applied else '基础生化段与二沉池出水'}<br/>"
            f"计算时间：{simulation.created_at:%Y-%m-%d %H:%M}",
            body_style,
        ),
        Paragraph("二、出水预测与达标结果", heading_style),
    ]
    rows = [["指标", "生化段出水", "最终出水", "去除率", "是否达标"]]
    effluent = simulation.effluent.model_dump()
    biological = (
        simulation.biological_effluent.model_dump()
        if simulation.biological_effluent is not None
        else effluent
    )
    removals = simulation.removal_rates.model_dump()
    compliance = simulation.compliance
    indicator_keys = [
        ("cod_mg_l", "cod"),
        ("nh4_n_mg_l", "nh4_n"),
        ("tn_mg_l", "tn"),
        ("tp_mg_l", "tp"),
        ("tss_mg_l", "tss"),
    ]
    for effluent_key, result_key in indicator_keys:
        rows.append(
            [
                INDICATOR_LABELS[effluent_key],
                f"{biological[effluent_key]:.3f}",
                f"{effluent[effluent_key]:.3f}",
                f"{removals[result_key] * 100:.1f}%",
                "达标" if compliance[result_key] else "超标",
            ]
        )
    table = Table(
        rows,
        colWidths=[36 * mm, 34 * mm, 34 * mm, 27 * mm, 24 * mm],
    )
    table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), "NotoSansSC"),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#dcefe8")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#174f43")),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#aebfba")),
                ("ALIGN", (1, 1), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("FONTSIZE", (0, 0), (-1, -1), 8.5),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f6f9f8")]),
            ]
        )
    )
    story.extend(
        [
            table,
            Paragraph("三、质量守恒与可信度检查", heading_style),
            Paragraph(
                f"水力相对误差：{simulation.mass_balance.hydraulic_relative_error:.3e}<br/>"
                f"化学需氧量回收比例：{simulation.mass_balance.cod_recovery:.3f}<br/>"
                f"总氮回收比例：{simulation.mass_balance.nitrogen_recovery:.3f}<br/>"
                f"质量守恒检查：{'通过' if simulation.mass_balance.passed else '未通过'}<br/>"
                f"稳态判定：{'达到' if simulation.convergence_reached else '尚未达到'}",
                body_style,
            ),
            Paragraph("四、假设与提示", heading_style),
        ]
    )
    for item in simulation.assumptions + simulation.warnings:
        story.append(Paragraph(f"• {item}", body_style))
    story.append(
        Paragraph(
            "本报告为模型计算结果。用于工程设计或运行决策前，必须使用独立时段实测数据完成校准和复核。",
            body_style,
        )
    )
    doc.build(story)


def _create_excel(
    path: Path, project: ProjectRecord, simulation: SimulationResult
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    workbook = Workbook()
    summary = workbook.active
    summary.title = "仿真结果"
    rows = [
        ("项目名称", project.name),
        ("污水厂名称", project.plant_name),
        ("主体工艺", project.process_type.value),
        ("计算引擎", simulation.engine),
        (
            "结果范围",
            "生化段加强化处理最终出水"
            if simulation.advanced_treatment_applied
            else "基础生化段与二沉池出水",
        ),
        ("仿真编号", simulation.simulation_id),
        ("计算时间", simulation.created_at.strftime("%Y-%m-%d %H:%M:%S")),
        (),
        ("指标", "生化段出水", "最终出水", "去除率", "达标情况"),
    ]
    effluent = simulation.effluent.model_dump()
    biological = (
        simulation.biological_effluent.model_dump()
        if simulation.biological_effluent is not None
        else effluent
    )
    removals = simulation.removal_rates.model_dump()
    for effluent_key, result_key in (
        ("cod_mg_l", "cod"),
        ("nh4_n_mg_l", "nh4_n"),
        ("tn_mg_l", "tn"),
        ("tp_mg_l", "tp"),
        ("tss_mg_l", "tss"),
    ):
        rows.append(
            (
                INDICATOR_LABELS[effluent_key],
                biological[effluent_key],
                effluent[effluent_key],
                removals[result_key],
                "达标" if simulation.compliance[result_key] else "超标",
            )
        )
    for row in rows:
        summary.append(row)
    summary.freeze_panes = "A10"
    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 42
    summary.column_dimensions["C"].width = 16
    summary.column_dimensions["D"].width = 16
    summary.column_dimensions["E"].width = 16
    for cell in summary[9]:
        cell.font = Font(bold=True, color="174F43")
        cell.fill = PatternFill("solid", fgColor="DCEFE8")

    mapping = workbook.create_sheet("组分映射与守恒")
    mapping.append(("模型组分", "浓度（毫克/升）"))
    for key, value in simulation.component_mapping.concentrations_mg_l.items():
        mapping.append((key, value))
    mapping.append(())
    mapping.append(("检查项", "结果"))
    mapping.append(("水力相对误差", simulation.mass_balance.hydraulic_relative_error))
    mapping.append(("化学需氧量回收比例", simulation.mass_balance.cod_recovery))
    mapping.append(("总氮回收比例", simulation.mass_balance.nitrogen_recovery))
    mapping.append(("总磷回收比例", simulation.mass_balance.phosphorus_recovery))
    mapping.append(("质量守恒是否通过", "是" if simulation.mass_balance.passed else "否"))
    mapping.column_dimensions["A"].width = 30
    mapping.column_dimensions["B"].width = 24
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="center")
    workbook.save(path)


def create_report(
    payload: ReportRequest,
    project: ProjectRecord,
    simulation: SimulationResult,
) -> ReportResult:
    title = payload.report_name or f"{project.name}工艺建模评估报告"
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    suffix = payload.report_format.value
    filename = f"{_safe_filename(project.name)}-{timestamp}.{suffix}"
    path = REPORT_DIR / filename
    if payload.report_format == ReportFormat.pdf:
        _create_pdf(path, title, project, simulation)
    else:
        _create_excel(path, project, simulation)
    return ReportResult(
        project_id=payload.project_id,
        status="ready",
        report_format=payload.report_format,
        message="报告已生成。",
        filename=filename,
        download_url=f"/api/reports/files/{filename}",
    )
