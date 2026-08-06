from collections import Counter
from datetime import date, datetime
from io import BytesIO

from app.models.schemas import (
    CalibrationImportResult,
    ModelCalibrationSample,
    ModelType,
    OperatingDataSource,
    PartialEffluentMeasurement,
    ProcessParameters,
    ProjectRecord,
    WaterQuality,
)


P_REMOVAL_PROCESS_VALUES = {
    "AAO",
    "SBR",
    "CASS",
    "UCT",
    "MUCT",
    "bardenpho5",
    "MBR",
    "IFAS",
}


def _clean_header(value: object) -> str:
    return str(value).replace("＊", "").strip() if value is not None else ""


def _records(sheet) -> list[dict[str, object]]:
    header_row = None
    headers: list[str] = []
    for row_number, row in enumerate(
        sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 10), values_only=True),
        start=1,
    ):
        candidate = [_clean_header(value) for value in row]
        if "工厂编号" in candidate and (
            "采样时间" in candidate or "记录时间" in candidate
        ):
            header_row = row_number
            headers = candidate
            break
    if header_row is None:
        raise ValueError(f"工作表“{sheet.title}”未找到工厂编号和时间字段表头。")
    return [
        dict(zip(headers, row, strict=False))
        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True)
        if any(value not in (None, "") for value in row)
    ]


def _number(value: object) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    return None


def _first_number(*values: object) -> float | None:
    for value in values:
        number = _number(value)
        if number is not None:
            return number
    return None


def _uncensored_number(
    row: dict[str, object], value_field: str, symbol_field: str
) -> float | None:
    symbol = str(row.get(symbol_field) or "").strip()
    if symbol in {"<", ">", "≤", "≥"}:
        return None
    return _number(row.get(value_field))


def _parse_datetime(value: object) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        try:
            return datetime.fromisoformat(text)
        except ValueError:
            return None
    return None


def _date_key(row: dict[str, object], date_field: str) -> tuple[str, str] | None:
    plant = str(row.get("工厂编号") or "").strip()
    value = _parse_datetime(row.get(date_field))
    if not plant or value is None:
        return None
    return plant, value.date().isoformat()


def _index_rows(
    rows: list[dict[str, object]], date_field: str
) -> tuple[dict[tuple[str, str], dict[str, object]], set[tuple[str, str]]]:
    grouped: dict[tuple[str, str], list[dict[str, object]]] = {}
    for row in rows:
        key = _date_key(row, date_field)
        if key is not None:
            grouped.setdefault(key, []).append(row)
    duplicates = {key for key, values in grouped.items() if len(values) > 1}
    return (
        {key: values[0] for key, values in grouped.items() if len(values) == 1},
        duplicates,
    )


def import_calibration_workbook(
    project: ProjectRecord, content: bytes
) -> CalibrationImportResult:
    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    required_sheets = {"进水数据", "出水数据", "运行参数"}
    missing_sheets = required_sheets.difference(workbook.sheetnames)
    if missing_sheets:
        raise ValueError(f"工作簿缺少工作表：{', '.join(sorted(missing_sheets))}")

    influent_rows = _records(workbook["进水数据"])
    effluent_rows = _records(workbook["出水数据"])
    operation_rows = _records(workbook["运行参数"])
    effluent_by_key, duplicate_effluent = _index_rows(
        effluent_rows, "采样时间"
    )
    operation_by_key, duplicate_operation = _index_rows(
        operation_rows, "记录时间"
    )
    samples = []
    skipped = 0
    skip_reasons: Counter[str] = Counter()
    for inlet in influent_rows:
        key = _date_key(inlet, "采样时间")
        if key is None:
            skipped += 1
            skip_reasons["工厂编号或采样时间无效"] += 1
            continue
        if key in duplicate_effluent:
            skipped += 1
            skip_reasons["同厂同日存在多条出水记录"] += 1
            continue
        if key not in effluent_by_key:
            skipped += 1
            skip_reasons["找不到同厂同日的出水记录"] += 1
            continue
        if key in duplicate_operation:
            skipped += 1
            skip_reasons["同厂同日存在多条运行记录，无法确定采用哪一条"] += 1
            continue
        if key not in operation_by_key:
            skipped += 1
            skip_reasons["找不到同厂同日的运行记录"] += 1
            continue
        outlet = effluent_by_key[key]
        operation = operation_by_key[key]
        inlet_values = {
            "flow_m3_d": _number(inlet.get("流量（立方米/日）")),
            "cod_mg_l": _uncensored_number(
                inlet,
                "化学需氧量（毫克/升）",
                "化学需氧量检出限符号",
            ),
            "bod_mg_l": _number(inlet.get("五日生化需氧量（毫克/升）")),
            "nh4_n_mg_l": _uncensored_number(
                inlet, "氨氮（毫克/升）", "氨氮检出限符号"
            ),
            "tn_mg_l": _number(inlet.get("总氮（毫克/升）")),
            "tp_mg_l": _number(inlet.get("总磷（毫克/升）")),
            "tss_mg_l": _number(inlet.get("悬浮物（毫克/升）")),
            "ph": _number(inlet.get("酸碱度")),
            "temperature_c": _number(inlet.get("水温（摄氏度）")),
        }
        required = (
            "flow_m3_d",
            "cod_mg_l",
            "nh4_n_mg_l",
            "tn_mg_l",
            "tp_mg_l",
            "tss_mg_l",
            "ph",
            "temperature_c",
        )
        missing_inlet = [name for name in required if inlet_values[name] is None]
        if missing_inlet:
            skipped += 1
            inlet_labels = {
                "flow_m3_d": "流量",
                "cod_mg_l": "化学需氧量",
                "nh4_n_mg_l": "氨氮",
                "tn_mg_l": "总氮",
                "tp_mg_l": "总磷",
                "tss_mg_l": "悬浮物",
                "ph": "酸碱度",
                "temperature_c": "水温",
            }
            skip_reasons[
                "缺少进水" + "、".join(inlet_labels[name] for name in missing_inlet)
            ] += 1
            continue
        measured_values = {
            "cod_mg_l": _uncensored_number(
                outlet,
                "化学需氧量（毫克/升）",
                "化学需氧量检出限符号",
            ),
            "nh4_n_mg_l": _uncensored_number(
                outlet, "氨氮（毫克/升）", "氨氮检出限符号"
            ),
            "tn_mg_l": _uncensored_number(
                outlet, "总氮（毫克/升）", "总氮检出限符号"
            ),
            "tp_mg_l": _uncensored_number(
                outlet, "总磷（毫克/升）", "总磷检出限符号"
            ),
            "tss_mg_l": _uncensored_number(
                outlet, "悬浮物（毫克/升）", "悬浮物检出限符号"
            ),
        }
        if not any(value is not None for value in measured_values.values()):
            skipped += 1
            skip_reasons["没有可校准的实测出水指标"] += 1
            continue
        operation_values = {
            "hrt_h": _number(operation.get("水力停留时间（小时）")),
            "srt_d": _number(operation.get("污泥龄（日）")),
            "internal_recycle_ratio": _number(operation.get("内回流比")),
            "sludge_recycle_ratio": _number(operation.get("污泥回流比")),
            "aerobic_do_mg_l": _number(
                operation.get("好氧池溶解氧（毫克/升）")
            ),
            "alkalinity_mg_l_caco3": _first_number(
                operation.get("碱度（毫克/升，以碳酸钙计）"),
                inlet.get("碱度（毫克/升，以碳酸钙计）"),
            ),
            "reactor_volume_m3": _first_number(
                operation.get("生化池总有效容积（立方米）"),
                operation.get("生化池有效容积（立方米）"),
            ),
            "clarifier_surface_area_m2": _number(
                operation.get("二沉池总表面积（平方米）")
            ),
            "clarifier_depth_m": _number(operation.get("二沉池有效水深（米）")),
            "settler_v_max_m_d": _number(
                operation.get("理论最大沉降速度（米/日）")
            ),
            "settler_v_max_practical_m_d": _number(
                operation.get("实用最大沉降速度（米/日）")
            ),
            "settler_tss_threshold_mg_l": _number(
                operation.get("沉降受阻临界污泥浓度（毫克/升）")
            ),
            "waste_sludge_flow_m3_d": _number(
                operation.get("实际排泥流量（立方米/日）")
            ),
            "mixed_liquor_tss_mg_l": _first_number(
                operation.get("池内污泥浓度（毫克/升）"),
                operation.get("混合液悬浮固体（毫克/升）"),
            ),
            "return_sludge_tss_mg_l": _number(
                operation.get("回流污泥浓度（毫克/升）")
            ),
            "waste_sludge_tss_mg_l": _number(
                operation.get("排泥污泥浓度（毫克/升）")
            ),
            "measured_total_energy_kwh_d": _number(
                operation.get("全厂同期实测总能耗（千瓦时/日）")
            ),
            "measured_dry_sludge_kg_d": _number(
                operation.get("全厂同期实测干污泥产量（千克/日）")
            ),
        }
        required_operation_names = {
            "hrt_h",
            "srt_d",
            "internal_recycle_ratio",
            "sludge_recycle_ratio",
            "aerobic_do_mg_l",
            "alkalinity_mg_l_caco3",
        }
        missing_operation = [
            name
            for name, value in operation_values.items()
            if name in required_operation_names and value is None
        ]
        if missing_operation:
            skipped += 1
            operation_labels = {
                "hrt_h": "水力停留时间",
                "srt_d": "污泥龄",
                "internal_recycle_ratio": "内回流比",
                "sludge_recycle_ratio": "污泥回流比",
                "aerobic_do_mg_l": "好氧池溶解氧",
                "alkalinity_mg_l_caco3": "碱度",
            }
            skip_reasons[
                "缺少运行参数"
                + "、".join(operation_labels[name] for name in missing_operation)
            ] += 1
            continue
        process_value = project.process_type.value
        optional_parameters = {
            name: value
            for name, value in operation_values.items()
            if name not in required_operation_names and value is not None
        }
        parameters = ProcessParameters(
            process_type=project.process_type,
            model_type=(
                ModelType.asm2d
                if process_value in P_REMOVAL_PROCESS_VALUES
                else ModelType.asm1
            ),
            hrt_h=operation_values["hrt_h"],
            srt_d=operation_values["srt_d"],
            internal_recycle_ratio=operation_values["internal_recycle_ratio"],
            sludge_recycle_ratio=operation_values["sludge_recycle_ratio"],
            aerobic_do_mg_l=operation_values["aerobic_do_mg_l"],
            aeration_power_kw=_number(operation.get("曝气功率（千瓦）") or 0)
            or 0,
            alkalinity_mg_l_caco3=operation_values["alkalinity_mg_l_caco3"],
            clarifier_solids_capture=(
                capture
                if (capture := _number(operation.get("二沉池固体捕集率")))
                is not None
                else 0.98
            ),
            **optional_parameters,
            operating_data_source=OperatingDataSource.measured,
        )
        samples.append(
            ModelCalibrationSample(
                group_id=key[0],
                sample_time=datetime.fromisoformat(key[1]),
                influent=WaterQuality(**inlet_values),
                measured=PartialEffluentMeasurement(**measured_values),
                parameters=parameters,
            )
        )
    warnings = [
        f"{count}行：{reason}。"
        for reason, count in skip_reasons.most_common()
    ]
    if not samples:
        warnings.append(
            "没有可用于完整模型校准的记录；不得用零值或模型默认值替代缺失实测数据。"
        )
    censored_influent_count = sum(
        str(row.get(symbol) or "").strip() in {"<", ">", "≤", "≥"}
        for row in influent_rows
        for symbol in ("化学需氧量检出限符号", "氨氮检出限符号")
    )
    censored_effluent_count = sum(
        str(row.get(symbol) or "").strip() in {"<", ">", "≤", "≥"}
        for row in effluent_rows
        for symbol in (
            "化学需氧量检出限符号",
            "氨氮检出限符号",
            "总氮检出限符号",
            "总磷检出限符号",
            "悬浮物检出限符号",
        )
    )
    if censored_influent_count or censored_effluent_count:
        warnings.append(
            "检出限符号数据未按精确值参与拟合："
            f"进水{censored_influent_count}项，出水{censored_effluent_count}项。"
        )
    coverage_fields = {
        "进水流量": (influent_rows, "流量（立方米/日）"),
        "进水化学需氧量": (influent_rows, "化学需氧量（毫克/升）"),
        "进水氨氮": (influent_rows, "氨氮（毫克/升）"),
        "进水总氮": (influent_rows, "总氮（毫克/升）"),
        "进水总磷": (influent_rows, "总磷（毫克/升）"),
        "进水悬浮物": (influent_rows, "悬浮物（毫克/升）"),
        "进水酸碱度": (influent_rows, "酸碱度"),
        "进水水温": (influent_rows, "水温（摄氏度）"),
        "出水化学需氧量": (effluent_rows, "化学需氧量（毫克/升）"),
        "出水氨氮": (effluent_rows, "氨氮（毫克/升）"),
        "出水总氮": (effluent_rows, "总氮（毫克/升）"),
        "出水总磷": (effluent_rows, "总磷（毫克/升）"),
        "水力停留时间": (operation_rows, "水力停留时间（小时）"),
        "污泥龄": (operation_rows, "污泥龄（日）"),
        "内回流比": (operation_rows, "内回流比"),
        "污泥回流比": (operation_rows, "污泥回流比"),
        "好氧池溶解氧": (operation_rows, "好氧池溶解氧（毫克/升）"),
        "碱度": (operation_rows, "碱度（毫克/升，以碳酸钙计）"),
        "生化池总有效容积": (operation_rows, "生化池总有效容积（立方米）"),
        "实际排泥流量": (operation_rows, "实际排泥流量（立方米/日）"),
        "池内污泥浓度": (operation_rows, "池内污泥浓度（毫克/升）"),
        "排泥污泥浓度": (operation_rows, "排泥污泥浓度（毫克/升）"),
        "二沉池总表面积": (operation_rows, "二沉池总表面积（平方米）"),
        "二沉池有效水深": (operation_rows, "二沉池有效水深（米）"),
    }
    field_coverage = {
        name: round(
            sum(_number(row.get(field)) is not None for row in rows)
            / max(len(rows), 1),
            4,
        )
        for name, (rows, field) in coverage_fields.items()
    }
    quality_score = round(
        sum(field_coverage.values()) / max(len(field_coverage), 1) * 100
    )
    duplicate_key_count = len(duplicate_effluent | duplicate_operation)
    recommendations = [
        f"补充{name}：当前有效覆盖率仅{coverage:.0%}。"
        for name, coverage in field_coverage.items()
        if coverage < 0.8
    ]
    if duplicate_key_count:
        recommendations.append(
            f"处理{duplicate_key_count}个同厂同日重复键，并明确并联线或采样批次。"
        )
    if len(samples) < 5:
        recommendations.append("每座污水厂至少形成五条完整配对记录，并保留独立验证时段。")
    if len(samples) >= 5 and quality_score >= 80 and not duplicate_key_count:
        readiness = "可进入校准"
    elif samples:
        readiness = "补充后校准"
    else:
        readiness = "不可校准"
    return CalibrationImportResult(
        project_id=project.id,
        imported_count=len(samples),
        skipped_count=skipped,
        groups=sorted({sample.group_id for sample in samples}),
        samples=samples,
        warnings=warnings,
        quality_score=quality_score,
        readiness=readiness,
        field_coverage=field_coverage,
        duplicate_key_count=duplicate_key_count,
        recommendations=recommendations,
    )
